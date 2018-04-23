import asyncio
import json
import logging
import sys

from aiohttp import ClientSession
from openpyxl import load_workbook
from sqlalchemy import Column, Integer, Float, String, DateTime, create_engine
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker

import utils

config = utils.config
loggers = [utils.def_log, utils.con_log]
Base = declarative_base()


class Monitoring(Base):

    __tablename__ = 'monitoring'

    id = Column('id', Integer, primary_key=True)
    ts = Column('timestamp', DateTime)
    url = Column('url', String, nullable=False)
    label = Column('label', String, nullable=False)
    response_time = Column('response_time', Float)
    status_code = Column('status_code', Integer)
    content_length = Column('content_length', Integer)


@utils.log_time(loggers)
def process_xlsx(filename):
    wb = load_workbook(filename=filename)
    ws = wb.active
    n = 1
    urls_to_fetch = []

    config = utils.load_config()

    while ws[f'{config["url_col"]}{n}'].value is not None:
        if ws[f'{config["fetch_col"]}{n}'].value == 1:
            urls_to_fetch.append((ws[f'{config["url_col"]}{n}'].value, ws[f'{config["label_col"]}{n}'].value))
        n += 1

    msg = f"Processed {sys.argv[1]}, found {len(urls_to_fetch)} urls to fetch."
    [logger.log(logging.INFO, msg) for logger in loggers]

    return urls_to_fetch


async def process_urls(session, urls, amount):
    main_queue = asyncio.Queue(maxsize=500)
    responses = []
    errors = []

    if len(urls) >= amount:
        urls_to_fetch = urls[:amount]
    else:
        urls_to_fetch = urls

    # we init the consumers, as the queues are empty at first,
    # they will be blocked on the main_queue.get()
    consumers = [asyncio.ensure_future(utils.consumer(main_queue=main_queue,
                                                      session=session,
                                                      responses=responses,
                                                      errors=errors))
                 for _ in range(amount)]

    await utils.producer(queue=main_queue, urls_to_fetch=urls_to_fetch)
    # wait for all items inside the main_queue to get task_done
    await main_queue.join()
    # cancel all coroutines
    for consumer_future in consumers:
        consumer_future.cancel()

    msg = f"Processed {len(urls_to_fetch)} urls, {len(errors)} errors total."
    [logger.log(logging.INFO, msg) for logger in loggers]

    process_errors(errors)
    return responses


@utils.log_time(loggers)
def process_errors(errors):
    try:

        with open(config['error_log_path'], 'r+') as j:
            try:
                msg = f"Trying to load data from error log({config['error_log_path']})"
                [logger.log(logging.DEBUG, msg) for logger in loggers]

                data = json.load(j)
                j.seek(0)
                j.truncate()
                data.get('entries').append(errors)
                json.dump(data, j, indent=4)

                msg = f"Successfully logged {len(errors)} errors."
                [logger.log(logging.INFO, msg) for logger in loggers]

            except json.JSONDecodeError:
                msg = f"Can't decode error log({config['error_log_path']}). Creating empty one..."
                [logger.log(logging.WARNING, msg) for logger in loggers]
                json.dump({'entries': errors}, j, indent=4)
    except FileNotFoundError:
        msg = "No error log found, creating new..."
        [logger.log(logging.DEBUG, msg) for logger in loggers]
        with open(config['error_log_path'], 'w'):
            pass
        process_errors(errors)


@utils.log_time(loggers)
def save_to_db(data_list):
    engine = create_engine(f'sqlite:///{config["db_path"]}')
    Base.metadata.create_all(bind=engine)
    Session = sessionmaker(bind=engine)
    db_session = Session()
    for each in data_list:
        monitoring = Monitoring(**each)
        db_session.add(monitoring)
    db_session.commit()
    db_session.close()

    msg = f"Saved {len(data_list)} entries to {config['db_path']}."
    [logger.log(logging.INFO, msg) for logger in loggers]


async def run(loop):
    try:
        amount = sys.argv[2]
    except IndexError:
        amount = config['urls_amount']

    urls_to_fetch = process_xlsx(sys.argv[1])

    async with ClientSession(loop=loop) as session:
        resps = await process_urls(session, urls=urls_to_fetch, amount=amount)

    save_to_db(resps)


@utils.log_time(loggers)
def main():
    loop = asyncio.get_event_loop()
    loop.run_until_complete(run(loop))

if __name__ == '__main__':
    main()
